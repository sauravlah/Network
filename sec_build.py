#!/usr/bin/python
#Name: sec_build.py
#Purpose: Takes a GIS Port matrix, and spits out a set of terraform files for SecLists
#Author: C.R.Chapman 
#Version: 0.1
###
###     TODO
###          - subnet aggregation
###          - port aggregation
###          - warn on duplicates
###          - warn on rule overlaps
###
###     data
###         - remove duplicates
###         - santize data to terraform requirements (ports, protocols, etc)
###
majorVersion = 0
minorVersion = 1
patchVersion = 'f'
#NOTES
#TODO

tenancy = "prod" #sandbox or prod
from openpyxl import Workbook, load_workbook
import re
import argparse
import os

import const

const.MAXRULE = 50
const.STARTROW = 0

class SecListRule:    #compartment, direction, statetype, Hostname, IP, Port, secListReference
    def __init__(self,SourceAdd,LowPort,HighPort,Protocol,Comment='',Stateful=True):
        #self.direction = Direction don't need as controled by the list, we choose the address (wrongly termed SourceAdd) on destination or source on ouotput
        self.sourceAdd = SourceAdd
        self.lowPort = LowPort
        self.highPort = HighPort
        self.protocol = Protocol
        self.stateful = Stateful
        self.comment = ''           #this comment will preceed

    def __str__(self):
        return self.outputRule()

    #this returns a string of the seclist rules
    def outputRule(self,direction=1):
        if self.protocol == "tcp": Protocol = "${var.Protocol-TCP}"
        elif self.protocol == "udp": Protocol = "${var.Protocol-UDP}"
        elif self.protocol == "icmp": Protocol = "${var.Protocol-ICMP}"
        elif self.protocol == "all": Protocol = "all"
        else: raise ValueError("Invalid protocol: {0} for rule with comment {1}".format(self.protocol,self.comment))    #do we need a better reference here to be able trackdown?
        
        if self.stateful: stateless = str("false")
        else: stateless = str("true")
        
        if direction: direction = "destination"
        else: direction = "source"

        if self.comment != '' : commentString = "###\n###\t"+self.comment+"\n###\n"
        else: commentString = ''
        if Protocol != "all": returnstring = commentString+"\t{\n\t"+self.protocol+"_options\t{\n\t\t\"max\"\t= \""+self.highPort+"\"\n\t\t\"min\"\t= \""+self.lowPort+"\"\n\t\t}\n\t\tstateless = \""+stateless+"\"\n\t\tprotocol = \""+Protocol+"\"\n\t\t"+direction+" = \""+self.sourceAdd+"\"\n\t},\n"
        else: returnstring = commentString+"\t{\n\tstateless = \""+stateless+"\"\n\tprotocol = \""+Protocol+"\"\n\t"+direction+" = \""+self.sourceAdd+"\"\n\t},\n"
        
        return returnstring

    def printRule(self):
        print self.__str__    

class SecList:
    #This is missing the compartment name
    def __init__(self,Name,DisplayName,Region):
        self.name = Name
        self.displayName = DisplayName
        self.region = Region
        self.ingressRules = list()
        self.egressRules = list()
        #ingress = 0, egress = 1

    def __contains__(other):
        if self.name == other.name and self.region == other.region:
            return True
        else:
            return False

    def __eq__(self, other):
        if self.name == other.name and self.region == other.region:
            return True
        else:
            return False

    def __str__(self):
        for rule in self.ingressRules:
            return rule
        for rule in self.egressRules:
            return rule
        
    def numOfIngress (self):
        return len(self.ingressRules)

    def numOfEgress (self):
        return len(self.egressRules)

    #creates new rule and adds to ingress lists
    def addIngressRule(self,Address,lowPort,upperPort,Protocol,Stateful = True):
        if self.numOfIngress() < const.MAXRULE:
            tempseclistrule = SecListRule(Address,lowPort,upperPort,Protocol,Comment='', Stateful = True)
            self.ingressRules.append(tempseclistrule)
        else: raise ValueError("Too many rules trying to be added to {0} Ingress".format(self.name))

    #creates new rule and adds to egress lists
    def addEgressRule(self,Address,lowPort,upperPort,Protocol,Stateful = True):
        if self.numOfEgress() < const.MAXRULE:
            tempseclistrule = SecListRule(Address,lowPort,upperPort,Protocol,Comment='',Stateful = True)
            self.egressRules.append(tempseclistrule)
        else: raise ValueError("Too many rules trying to be added to {0} Egress".format(self.name))

    #output both SecList
    def outputSecList(self): #returns a list, one line per slot
        alist = str()
        #add header section
        alist = "resource \"baremetal_core_security_list\" \""+self.name+"\" {\n\tcompartment_id\t= \"${var.prod_compartment_git_networks_ocid}\"\n\tdisplay_name\t= \""+self.displayName+"\"\n\tvcn_id\t\t\t= \"${baremetal_core_virtual_network.oragit-"+self.region+"-vcn1.id}\"\n"
        #add ingress header
        alist = alist + "ingress_security_rules = [\n"
        #ingress
        for line in self.ingressRules:
            alist = alist + line.outputRule(0)
        #add ingress footer
        alist = alist + "\t]\n"
        #add egress header
        alist = alist + "egress_security_rules = [\n"
        #egress
        for line in self.egressRules:
            alist = alist + line.outputRule(1)
        #add egress footer
        alist = alist + "\t]\n"
        #add footer section
        alist = alist + "}"
        return alist

    def getName(self):
        return self.name

    def printSecList(self):
        print self.__str__()

class Subnet:
    def __init__(self, Name, DisplayName,Region,CIDR,Compartment,Environment):
        self.name = Name
        self.displayName = DisplayName
        self.region = Region
        self.cidr = CIDR
        self.environment = Environment
        self.compartment = Compartment
        self.subnetName = "oragit-"+Compartment+"-sec-vcn1-"+Environment

        self.seclists = [ SecList(convertTosec(Name)+'-1', convertTosec(Name)+'-1', Region), SecList(convertTosec(Name)+'-2', convertTosec(Name)+'-2', Region), SecList(convertTosec(Name)+'-3', convertTosec(Name)+'-3', Region)]

    def __contains__(other):
        if self.name == other.name and self.region == other.region:
            return True
        else:
            return False

    def __eq__(self, other):
        if isinstance(other, Subnet):
            return self.name == other.name and self.region == other.region
        else:
            return False

    def __hash__(self):
        return hash(self.name)

    #this needs to be sorted out for the future
    #def __repr__(self):
    #    return '%s(%r)' % (self.__class__.__name__,self.name,self.displayName,self.region,self.environment,self.compartment,self.subnetName,self.seclists)
    
    def __str__(self):
        return str(self.displayName)

    #Add a new list to a list dependent on the direction given (0 ingress, 1 egress
    def addSecListLine(self, Direction, Address, lowPort, upperPort, Protocol, Stateful = True):
        #try:
            if Direction:
                if self.seclists[0].numOfEgress() < const.MAXRULE:
                    self.seclists[0].addEgressRule(Address, lowPort, upperPort, Protocol, Stateful)
                elif self.seclists[1].numOfEgress() < const.MAXRULE:
                    self.seclists[1].addEgressRule(Address, lowPort, upperPort, Protocol, Stateful)
                elif self.seclists[2].numOfEgress() < const.MAXRULE:
                    self.seclists[2].addEgressRule(Address, lowPort, upperPort, Protocol, Stateful)
                else:
                     raise SystemExit("Too many Egress rules on {0}".format(self.name))

            else:
                if self.seclists[0].numOfIngress() < const.MAXRULE:
                    self.seclists[0].addIngressRule(Address, lowPort, upperPort, Protocol, Stateful)
                elif self.seclists[1].numOfIngress() < const.MAXRULE:
                    self.seclists[1].addIngressRule(Address, lowPort, upperPort, Protocol, Stateful)
                elif self.seclists[2].numOfIngress() < const.MAXRULE:
                    self.seclists[2].addIngressRule(Address, lowPort, upperPort, Protocol, Stateful)
                else:
                    raise SystemExit("Too many Ingress rules on {0}".format(self.name))

    def getName(self):
        return self.name

    def getSecList(self, i):
        return self.seclists[i]

    def matchThisSubnet(self, subnetrange):
        #eventually need this to check if a subnet range contains this subnet
        return False

    #def returnSecLists(self):
    #    secList = list()
    #    self.seclists[0].outputSecList()
    #    print "From Subnet"
    #    return secList

    #prints out number of rules totally used by the applications
    def rulesUsed(self):
        egress = self.seclists[0].numOfEgress() + self.seclists[1].numOfEgress() + self.seclists[2].numOfEgress()
        ingress = self.seclists[0].numOfIngress() + self.seclists[1].numOfIngress() + self.seclists[2].numOfIngress()
        print "Egress rules used " + str(egress)
        print "Ingress rules used " + str(ingress)

    #outputs all seclitst
    def outputSecLists(self):
        returnString = str()
        for internalList in self.seclists:
            returnString = returnString + internalList.outputSecList()
        return returnString

    #output a single seclist (egress and ingress, takes a decimal)
    def outputSecList(self, seclist):
        returnString = str()
        return self.seclists[seclist].outputSecList()
#END OF Subnet

class SubnetFile:
    def __init__(self, filename, startrow, subnetfilename, subnetsheet):
        try:
            #if we've defined this then we must have some data else forget it
            if subnetfilename is not None:
                self.ws2 = load_workbook(subnetfilename) #throw an error is unable to open the file
        except IOError as e:
            print "Unable to open the Subnet file {0}: {1}".format(subnetfilename, e)
            exit(0)
        except BadZipfile as e:
            print "File corrupted {0}".format(e)
        else:
            self.SNSheet = self.ws2[subnetsheet]

    #takes a dictionary and adds subnets to it
    def addSubnets(self, subnets):
        for index, row in enumerate(self.getSubnets()):
            #finally we add a special case for the general seclists (10.15.0.0/16 /17 or /18)
            pass

    def getSubnets(self):
        alist = list()
        #iterate through all the rows
        for index, row in enumerate(self.SNSheet.iter_rows()):
            #format is name, ad number, vcn, region, compartment, tenancy,
            #def __init__(self, Name, DisplayName,Region,CIDR,Compartment,VCN)
            if not( row[0].internal_value == None or row[1].internal_value == None or row[2].internal_value == None or row[3].internal_value == None or row[4].internal_value == None or row[5].internal_value == None):
                addresses = str(row[1].internal_value).split('\n') + str(row[3].internal_value).split('\n')
                for line in addresses:
                    if not ipFormatCheck(line.strip()):
                        raise ValueError("Invalid IP address on line "+str(index+1)+" "+line) 
                if not isPortRange(row[5].internal_value):
                    raise ValueError("Invalid port range on line "+str(index+1))
                alist.append(row)
            elif index > self.startRow and not (row[0].internal_value == None or row[1].internal_value == None or row[2].internal_value == None or row[3].internal_value == None or row[4].internal_value == None or row[5].internal_value == None):
                    raise ValueError("Invalid data on row "+str(index+1))
        return alist
    

class PortMatrix:
    def __init__(self, filename, startrow, pmsheet):
        try:
            self.ws = load_workbook(filename) #throw an error is unable to open the file
        except IOError as e:
            print "Unable to open the Port Matrix file {0}: {1}".format(filename, e)
            exit(0)
        except BadZipfile as e:
            print "File corrupted {0}".format(e)
            exit(0)
        else:
            self.currentRow = 0
            self.startRow = startrow
            self.PMSheet = self.ws[pmsheet]

    #set of utility functions for crafting elements
    def _isPortRange(stringIN):
        stringF = str(stringIN).strip()
        #is the cell containing either 1 number or two number separated by a dash or Any:Any
        if re.search('^\d{1,5}(-\d{1,5})?$',stringF) or re.search('^all$',stringF):
            return True
        raise ValueError("Invalid Port {0}".format(stringIN))
        return False

    #takes a dictionary for subnets and adds rules to
    #TODO reduce the 
    def addToConfig(self, subnets, seclists):
        if not isinstance(subnets, dict):
            raise ValueError("{0} is not a valid".format(subnets))
        if not isinstance(seclists, dict):
            raise ValueError("{0} is not a valid".format(seclists))
        for index, row in enumerate(self.getAllRows()):
            count = const.STARTROW
            rowindex = self.startRow
            #check we have a valid port range or skip
            cellF = str(row[5].internal_value).strip()
            count = 0
            if self._isPortRange(cellF): #TODO should this happen or should we just bomb out? Should we look at more parameters
                count = count + 1
                #strip the input to line to a single first line to see if we want to create 
                cellA = re.search(r'^(.*)$',str(row[0].internal_value),re.M).group(0)
                cellC = re.search(r'^(.*)$',str(row[2].internal_value),re.M).group(0)

                #if there are oragit subnets in the source create an egress rule for source subnet
                if re.search('^oragit-(phx|ash)1-net-vcn\d-ad[1-3]-.*$',cellA): #TODO this would probably be best testing against a list
                    #mash to the ad1 subnet
                    subnet1 = mashSubnetName(cellA).strip()
                    #print out where we got an incorrect 
                    if subnet1 == None: print index+rowindex, ":", subnet1
                    
                    #TODO is this subnet already made, if not make it? We shouldn't be doing this. Our subnets should already be created
                    if subnet1 not in subnets:
                        subnets[subnet1] = Subnet(subnet1, subnet1,returnRegionFromName(subnet1),"Dummy","Dummy","Dummy")
                    
                    #add new rule/s to subnet (dependent on the number of lines in the opposite B or D column
                    #because cells can have multiple rows
                    lineint = 0
                    #getting the oragit* cidr reference or an external address range
                    cellD = returnAddressesFromString(str(row[3].internal_value))
                    for line in cellD:
                        cellE = str(row[4].internal_value)
                        minPort, maxPort = returnPorts(cellF)
                        if minPort == 'all': cellE = 'all'
                        elif re.search('udp',cellE,re.I): cellE = 'udp'
                        else: cellE = 'tcp'
                        Protocol = cellE
                        address = str(nameOrAddress(str(row[2].internal_value).strip(),str(row[3].internal_value).strip(), lineint)).strip()
                        addresult = subnets[subnet1].addSecListLine(1, address.strip(), minPort, maxPort, Protocol, True)
                        if addresult ==0:
                            print "Found duplication at line {0}".format(index+rowindex+2)
                        lineint = lineint + 1

                #if there is an oragit rule in the destination create and ingress rule for the destination subnet
                if re.search('^oragit-(phx|ash)1-net-vcn\d-ad[1-3]-.*$',cellC): #TODO this would probably be best testing against a list
                    #mash to the ad1 subnet
                    subnet1 = mashSubnetName(cellC).strip()
                    if subnet1 == None: print index+rowindex, ":", subnet1
                    
                    #TODO is this subnet already made, if not make it? We shouldn't be doing this. Our subnets should already be created
                    if subnet1 not in subnets:
                        subnets[subnet1] = Subnet(subnet1, subnet1,returnRegionFromName(subnet1),"Dummy","Dummy","Dummy")
                    
                    #add new rule/s to subnet (dependent on the number of lines in the opposite B or D column
                    lineint = 0
                    cellB = returnAddressesFromString(str(row[1].internal_value))
                    for line in cellB:
                        cellE = str(row[4].internal_value)
                        minPort, maxPort = returnPorts(cellF)
                        if minPort == 'all': cellE = 'all'
                        elif re.search('udp',cellE,re.I): cellE = 'udp'
                        else: cellE = 'tcp'
                        Protocol = cellE
                        address = nameOrAddress(str(row[0].internal_value).strip(),str(row[1].internal_value).strip(), lineint)
                        addresult = subnets[subnet1].addSecListLine(0, address.strip(), minPort, maxPort, Protocol, True)
                        if addresult ==0:
                            print "Found duplication at line {0}".format(index+rowindex+2)
                        lineint = lineint + 1
                #else we skip, implicit

    def getAllRows(self):
        alist = list()
        version = self.getVersion()
        for index, row in enumerate(self.PMSheet.iter_rows()):
            if index > self.startRow:
                if not( row[0].internal_value == None or row[1].internal_value == None or row[2].internal_value == None or row[3].internal_value == None or row[4].internal_value == None or row[5].internal_value == None):
                    addresses = row[1].internal_value.encode('ascii',errors='ignore').split('\n') + row[3].internal_value.encode('ascii',errors='ignore').split('\n')
                    for line in addresses:
                        if not ipFormatCheck(line.strip()):
                            raise ValueError("Invalid IP address on line "+str(index+1)+" "+line) 
                    if not isPortRange(row[5].internal_value):
                        raise ValueError("Invalid port range on line "+str(index+1))
                    alist.append(row)
                elif index > self.startRow and not (row[0].internal_value == None or row[1].internal_value == None or row[2].internal_value == None or row[3].internal_value == None or row[4].internal_value == None or row[5].internal_value == None):
                        raise ValueError("Invalid data on row "+str(index+1))
            #else:
            #    for i in range(0,6):
            #        print row[i].internal_value,
            #    print
        return alist

    def getNextRow(self):
        alist = list()
        self.currentRow = self.currentRow + 1
        for i in 'abcdef':
            cell = i+int(self.currentRow)
            alist.append = self.PMSheet[cell].internal_value
        return alist

    def getCell(self, cell):
        if re.search('\w+\d+'): return self.PMSheet[cell]
        else: return None

    def getChanges(self):
        returnstring = list()
        version = self.getVersion()
        for index, row in enumerate(self.PMSheet.iter_rows()):
            if re.search(str(version),str(row[7].internal_value)):
                #print index+1,
                #for i in range(0,5):
                #    print str(row[i].internal_value).split(),
                #print
                templist = list()
                templist.append(index+1)
                for i in range(0,5):
                    templist.append(str(row[i].internal_value).split())
                returnstring.append(templist)
        return returnstring

    def getVersion(self):
        return self.PMSheet['E5'].internal_value
#END OF PortMatrix

#set of utillity functions for crafting elements
def isPortRange(stringIN):
    stringF = str(stringIN).strip()
    #is the cell containing either 1 number or two number separated by a dash or Any:Any
    if re.search('^\d{1,5}(-\d{1,5})?$',stringF) or stringF == 'all':
        return True
    raise ValueError("Invalid Port {0}".format(stringIN))
    return False

#detects an oragitnet address
def isORAGITnet(string):
    string = string.strip()
    if re.search('^oragit-(ash|phx)1-net-vcn1-ad[1-3]-prod-ngcc-*$',string):
        return True
    return False

#net to cidr address
def convertTocidr(string):
    match = re.match('^(oragit-)(ash|phx)(\d-)net-(vcn1-ad[1-3]-.*)$',string)
    return match.group(1)+match.group(2)+match.group(3)+"cidr-"+match.group(4)

#net to sec value
def convertTosec(string):
    match = re.match('^(oragit-)(ash|phx)(\d-)net-(vcn1-)ad[1-3]-(.*)$',string)
    return match.group(1)+match.group(2)+match.group(3)+"sec-"+match.group(4)+match.group(5)

#converts all adX addresses to AD1
def mashSubnetName(subnetName):
    matches = re.search('^(oragit-)(ash|phx)(\d-net-vcn1-ad)[1-3](.*)$', subnetName)
    if matches:
        return matches.group(1)+matches.group(2)+matches.group(3)+'1'+matches.group(4)
    else:
        return None

#from a net address, returns the region (phoenix or ashburn)
def returnRegionFromName(subnetName):
    matches = re.search('^oragit-(ash|phx)(\d+)-net-vcn1-ad[1-3].*$', subnetName)
    if matches:
        return matches.group(1)+matches.group(2)
    else:
        return None

#getting a multiline value and responding (this is due to excel spreadsheet cells with multiline addresses and networks)
def returnAddressesFromString(subnetString):
    return subnetString.split('\n')

#figure out whether a range, a single port, or all ports
def returnPorts(portString):
    #so no parsing of the actual values here
    portString = str(portString).strip()
    if '-' in portString: return portString.split('-')
    elif portString == 'all':
        return ('all','all')
    else: return (portString, portString)

#This will figure out which of and address or ip address to use. If the subnet name starts with oragit it presumes use the name
#It also chose which which pair to look at depending on the line value, which needs to be calculated by the calling code
#If the line exceeds the number of lines in the corresponding choice, it defaults to the first one
#finally if a it is an IP address it makes sure there is a valid CIDR suffix attached, otherwise sets it to /32
def nameOrAddress(cell1,cell2, line):
    names = re.split("\n|,",cell1)
    addresses = re.split("\n|,",cell2)
    #some lines have one name and two (or more) addresses, trying do detect which one to use
    if line >= len(names) or line >= len(addresses): line = 0 #default to the first line
    if re.search("^oragit",names[line].strip()): return "${var."+convertTocidr(str(names[line]).strip())+"}"
    else:
        addressreturn = str(addresses[line])
        #see if it ends in /XX
        m = re.search('^([\d.]+)/(\d\d?)$',addressreturn)
        if m:
            #if yes is the value between 0 and 32 inclusive
            ##      fixed bug < to > MAB 
            if m.group(2) > -1 or m.group(2) < 33:
                return addressreturn
        #else make it /32 as a safety precaution
        m = re.search('^([\d.]+)',addressreturn)
        if m:
            return m.group(1)+'/32'
        else: return 'error'


def ipFormatCheck(ip_str):
   pattern = r"\b(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\b"
   if re.match(pattern, ip_str):
      return True
   else:
      return False


if __name__ == '__main__':
    
    parser = argparse.ArgumentParser(description='Creates blank subnet and seclist files')
    parser.add_argument("-s", help="Excel Port Matrix:Work Sheet", metavar="<seclist spreadsheet>", required="True")
    parser.add_argument("-n", help="Excel Port Matrix Network file,", metavar="<Subnet List spreadsheet>")
    parser.add_argument("-v", help="Prints out versioning information", action='store_true', dest="version" ,default="False")
    parser.add_argument("-l", help="Prints out number of seclist rules used", action='store_true', dest="lines" ,default="False")
    args = parser.parse_args()

    matchstr = re.search(r"^(.*):(.*)$",args.s)
    portmatrixfile = matchstr.group(1)
    portmatrixsheet = matchstr.group(2)
    print portmatrixfile
    print portmatrixsheet
    #spreadsheet = PortMatrix( args.s, 6, "SIT PM to impl")
    #spreadsheet = PortMatrix( args.s, 6, "PRD PM- Interim  SIT impl")
    spreadsheet = PortMatrix( portmatrixfile, 6, portmatrixsheet)
    

    if args.version == True:
        print ("sec Build Version: {0}.{1}.{2}".format(majorVersion,minorVersion,patchVersion))
        print "Port Matrix Vesion: {0}".format(spreadsheet.getVersion())
        exit(0)
    
    #strwb = load_workbook(portmatrixfile)

    #list of subnets
    subnets = dict()

    count = const.STARTROW
    rowindex = 6
    
    for index, row in enumerate(spreadsheet.getAllRows()):
        cellF = str(row[5].internal_value).strip()
        #print "Ports {0} {1}".format(rowindex,cellF)
        if isPortRange(cellF):
            count = count + 1
            #strip the input to line to a single first line
            cellA = re.search(r'^(.*)$',str(row[0].internal_value.encode('ascii', errors='ignore')),re.M).group(0)
            cellC = re.search(r'^(.*)$',str(row[2].internal_value.encode('ascii', errors='ignore')),re.M).group(0)

            #if there are oragit subnets in the source create an egress rule for source subnet
            if re.search('^oragit-(phx|ash)1-net-vcn\d-ad[1-3]-.*$',cellA): #TODO this would probably be best testing against a list
                #mash to the ad1 subnet
                subnet1 = mashSubnetName(cellA).strip()
                if subnet1 == None: print rowindex, ":", subnet1
                
                #is this subnet already made, if not make it?
                if subnet1 not in subnets:
                    subnets[subnet1] = Subnet(subnet1, subnet1,returnRegionFromName(subnet1),"Dummy","Dummy","Dummy")
                
                #add new rule/s to subnet (dependent on the number of lines in the opposite B or D column
                lineint = 0
                cellD = returnAddressesFromString(str(row[3].internal_value.encode('ascii', errors='ignore')))
                for line in cellD:
                    cellE = str(row[4].internal_value)
                    minPort, maxPort = returnPorts(cellF)
                    if minPort == 'all': cellE = 'all'
                    elif re.search('udp',cellE,re.I): cellE = 'udp'
                    else: cellE = 'tcp'
                    Protocol = cellE
                    address = str(nameOrAddress(str(row[2].internal_value.encode('ascii', errors='ignore')).strip(),str(row[3].internal_value.encode('ascii', errors='ignore')).strip(), lineint)).strip()
                    subnets[subnet1].addSecListLine(1, address.strip(), minPort, maxPort, Protocol, True)
                    lineint = lineint + 1

            #if there is an oragit rule in the destination create and ingress rule for the destination subnet
            if re.search('^oragit-(phx|ash)1-net-vcn\d-ad[1-3]-.*$',cellC): #TODO this would probably be best testing against a list
                #mash to the ad1 subnet
                subnet1 = mashSubnetName(cellC).strip()
                if subnet1 == None: print rowindex, ":", subnet1
                
                #is this subnet already made, if not make it?
                if subnet1 not in subnets:
                    subnets[subnet1] = Subnet(subnet1, subnet1,returnRegionFromName(subnet1),"Dummy","Dummy","Dummy")
                
                #add new rule/s to subnet (dependent on the number of lines in the opposite B or D column
                lineint = 0
                cellB = returnAddressesFromString(row[1].internal_value.encode('ascii', errors='ignore'))
                for line in cellB:
                    cellE = str(row[4].internal_value)
                    minPort, maxPort = returnPorts(cellF)
                    if minPort == 'all': cellE = 'all'
                    elif re.search('udp',cellE,re.I): cellE = 'udp'
                    else: cellE = 'tcp'
                    Protocol = cellE
                    address = nameOrAddress(str(row[0].internal_value.encode('ascii', errors='ignore')).strip(),str(row[1].internal_value.encode('ascii', errors='ignore')).strip(), lineint)
                    subnets[subnet1].addSecListLine(0, address.strip(), minPort, maxPort, Protocol, True)
                    lineint = lineint + 1
            #else we skip, implicit
    
    
    for key in subnets:
        if isinstance(subnets[key], Subnet):
            print "*Start of {0}".format(subnets[key].getName())
            #print subnets[key]
            print subnets[key].rulesUsed()
            #print subnets[key].outputSecLists()
            for i in range(0,3):
                seclistname = subnets[key].getSecList(i)
                print "Writing: " +seclistname.getName()
                ofile = open(seclistname.getName()+".tf",'w')
                ofile.write(seclistname.outputSecList()+"\n")
                ofile.close()
                #if seclistname.getName() == "oragit-ash1-sec-vcn1-prod-ngcc-external-mt1":
                #    print seclistname.outputSecList()
            print "*End of {0}".format(subnets[key].getName())
    print "Number of rows found: "+str(count-const.STARTROW)
